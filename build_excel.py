import sys, json, base64, io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU, cm_to_EMU

# =========================
# LOAD INPUT DATA
# =========================
# (Keeping your sys.argv logic, but using defaults for demonstration)
try:
    data = json.loads(open(sys.argv[1]).read())
except:
    data = {
        'staffName': 'JOHN DOE', 
        'staffUnit': 'Operations', 
        'tasks': [], 
        'excelFile': 'Daily_Tracker.xlsx'
    }

staff_name = data['staffName']
staff_unit = data['staffUnit']
tasks      = data['tasks']
out_path   = data['excelFile']

# =========================
# EXACT IMAGE COLORS (Hex)
# =========================
GREEN_DARK   = "385623" # The "Daily Deliverables Tracker" green
GREEN_LIGHT  = "E2EFDA" # Instruction box green
GREEN_LABEL  = "548235" # Name/Unit label green
YELLOW_BOX   = "FFFF00"
WHITE        = "FFFFFF"
BLACK        = "000000"
HEADER_GREY  = "D9D9D9"

# =========================
# HELPERS
# =========================
def fill(color):
    return PatternFill("solid", fgColor=color)

def font(size=11, bold=False, color=BLACK, name="Calibri"):
    return Font(name=name, size=size, bold=bold, color=color)

thin = Side(style='thin', color=BLACK)
def get_border():
    return Border(left=thin, right=thin, top=thin, bottom=thin)

# =========================
# WORKBOOK SETUP
# =========================
wb = Workbook()
ws = wb.active
ws.title = "Daily Tracker"
ws.sheet_view.showGridLines = False # Cleaner look like the image

# Column Widths to match the proportions
widths = {"A": 6, "B": 25, "C": 45, "D": 12, "E": 18, "F": 40, "G": 40}
for col, width in widths.items():
    ws.column_dimensions[col].width = width

# =========================
# HEADER ROW 1: Instructions | Logo | Tracker Info
# =========================
ws.row_dimensions[1].height = 90

# B1: Yellow Meeting Box
ws["B1"] = "Team Stand Up Meeting:\nMorning - 9:00am\nProgress Check-in - 4:00pm"
ws["B1"].fill, ws["B1"].font, ws["B1"].border = fill(YELLOW_BOX), font(size=11, bold=True), get_border()
ws["B1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# C1: Placeholder for Logo (Image anchor is handled later)
# Instructions Box (Merged F1:G1)
ws.merge_cells("F1:G1")
ws["F1"] = ("How to Complete the Tracker:\n"
            "Morning 8:00am: Fill your Proposed Tasks + Expected Completion Time\n\n"
            "Evening (At The End Of The Day) 4:00pm: Fill in your completed task and "
            "share with your line manager/HOD and upload necessary proof to a shared drive")
ws["F1"].fill = fill(GREEN_LIGHT)
ws["F1"].font = font(size=10, bold=True)
ws["F1"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
for cell in ws["F1:G1"][0]: cell.border = get_border()

# =========================
# NAME & UNIT ROWS (2 & 3)
# =========================
ws.row_dimensions[2].height = 25
ws.row_dimensions[3].height = 25

# Labels (A2, A3)
for addr, txt in [("A2", "Name"), ("A3", "Unit")]:
    ws[addr] = txt
    ws[addr].fill, ws[addr].font, ws[addr].border = fill(GREEN_LABEL), font(size=12, bold=True, color=WHITE), get_border()
    ws[addr].alignment = Alignment(horizontal="center", vertical="center")

# Values (B2, B3)
ws["B2"], ws["B3"] = staff_name.upper(), staff_unit
for addr in ["B2", "B3"]:
    ws[addr].fill, ws[addr].font, ws[addr].border = fill(GREEN_DARK), font(size=12, bold=True, color=WHITE), get_border()
    ws[addr].alignment = Alignment(horizontal="center", vertical="center")

# Main Title (C2:F3 Merged)
ws.merge_cells("C2:F3")
ws["C2"] = "Daily Deliverables Tracker"
ws["C2"].fill = fill(GREEN_DARK)
ws["C2"].font = font(size=24, bold=True, color=WHITE, name="Calibri Light")
ws["C2"].alignment = Alignment(horizontal="center", vertical="center")
# Apply borders/fill to merged range
for r in range(2, 4):
    for c in range(3, 7):
        ws.cell(row=r, column=c).border = get_border()
        ws.cell(row=r, column=c).fill = fill(GREEN_DARK)

# Date and Time Boxes (G2, G3)
current_dt = datetime.now()
ws["G2"], ws["G3"] = f"Date: {current_dt.strftime('%d/%m/%Y')}", f"Time: {current_dt.strftime('%I:%M %p')}"
for addr in ["G2", "G3"]:
    ws[addr].fill, ws[addr].font, ws[addr].border = fill(GREEN_DARK), font(size=11, bold=True, color=WHITE), get_border()
    ws[addr].alignment = Alignment(horizontal="left", vertical="center")

# =========================
# TABLE HEADERS (Row 4)
# =========================
ws.row_dimensions[4].height = 45
headers = ["S/N", "Clients/Focus Area", "Proposed Task/Deliverable for the Day", 
           "Priority\n(H/M/L)", "Expected Time\nof Completion", 
           "Actual Deliverables for Today", "Achievement/Result for the Day"]

for i, text in enumerate(headers, start=1):
    cell = ws.cell(row=4, column=i)
    cell.value = text
    cell.fill, cell.font, cell.border = fill(HEADER_GREY), font(size=10, bold=True), get_border()
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# =========================
# TASK ROWS
# =========================
start_row = 5
if not tasks: # Add empty rows for visual consistency if no data
    tasks = [{"client":""}] * 5

for idx, task in enumerate(tasks, start=1):
    row = start_row + idx - 1
    ws.row_dimensions[row].height = 60
    
    vals = [idx, task.get("client",""), task.get("proposed",""), task.get("priority",""), 
            task.get("time",""), task.get("actual",""), task.get("result","")]
    
    for col_idx, val in enumerate(vals, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = val
        cell.border = get_border()
        cell.alignment = Alignment(horizontal="center" if col_idx in [1, 4, 5] else "left", 
                                   vertical="top", wrap_text=True)

# =========================
# LOGO PLACEMENT (CENTERED)
# =========================
try:
    LOGO_B64 = "/9j/4AAQSkZJRgABAQEA3ADcAAD/2wBDAAIBAQEBAQIBAQECAgICAgQDAgICAgUEBAMEBgUGBgYFBgYGBwkIBgcJBwYGCAsICQoKCgoKBggLDAsKDAkKCgr/2wBDAQICAgICAgUDAwUKBwYHCgoKCgoKCgoKCgoKCgoKCgoKCgoKCgoKCgoKCgoKCgoKCgoKCgoKCgoKCgoKCgoKCgr/wAARCAG0AgkDASIAAhEBAxEB/8QAHwAAAQUBAQEBAQEAAAAAAAAAAAECAwQFBgcICQoL/8QAtRAAAgEDAwIEAwUFBAQAAAF9AQIDAAQRBRIhMUEGE1FhByJxFDKBkaEII0KxwRVS0fAkM2JyggkKFhcYGRolJicoKSo0NTY3ODk6Q0RFRkdISUpTVFVWV1hZWmNkZWZnaGlqc3R1dnd4eXqDhIWGh4iJipKTlJWWl5iZmqKjpKWmp6ipqrKztLW2t7i5usLDxMXGx8jJytLT1NXW19jZ2uHi4+Tl5ufo6erx8vP09fb3+Pn6/8QAHwEAAwEBAQEBAQEBAQAAAAAAAAECAwQFBgcICQoL/8QAtREAAgECBAQDBAcFBAQAAQJ3AAECAxEEBSExBhJBUQdhcRMiMoEIFEKRobHBCSMzUvAVYnLRChYkNOEl8RcYGRomJygpKjU2Nzg5OkNERUZHSElKU1RVVldYWVpjZGVmZ2hpanN0dXZ3eHl6goOEhYaHiImKkpOUlZaXmJmaoqOkpaanqKmqsrO0tba3uLm6wsPExcbHyMnK0tPU1dbX2Nna4uPk5ebn6Onq8vP09fb3+Pn6/9oADAMBAAIRAxEAPwD9/KKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiimmRB3oAdRTfNj/vUech6GgB1FGRRketABRRnHWigAoozRketABRTTJjtRvFADqKb5gAyaFlRjgGgB1FAYN0NFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAVDM4jXeey5qas3xDeLZaTcXbdI4GY/lWGIqexoyn2TC9tTym+/bG+G+j+LbvwxrUVxA1rcNEZwu5WIrvPCPxc+H/jZFk8O+J7Wfd/yzEgDZ9MGvgXxPfNrPiK+1R2+a4vZJBn0LH+lV7C+1HS7hbzTb2SGReQ0bEGv52peL2ZYPMJ061NSgpNK29rnmfXnGVmfpMLmM42n8alVgOGr4r+GH7Yfj3wW8dn4kk/tSzXhvMP7xR7Gvp/4X/GrwX8VtOW68PakplXHnWrt86H0I/rX61w7xzkvEEVGnPlm+jOyniKdTZnbSU6ohMHO6niRSM19pGVzcV/u0xGXP3qVpFK8Gq09yiJlpAoHViaroLrdk0jAHg8VVvtUs7K2a6u7hI41GXZ3CqB1618iftt/8Fg/2d/2TBdeFNE1GPxR4qQEf2Zp8waOBv+mjjgfQc1+UP7VP/BVf9q/9qO8ntNS8az6Lo0j/ALvSdJkMabfRiOWrlqYmnFWPms04mwGX3Sd5eR+xH7R//BWH9jj9nD7RpuufEi31fVoODpeisJ2DYzgsPlH5mvXf2Yvj34V/af8AgvoPxw8GQyQ6fr1s00MMzAvHtdkKnHfKmv5hbmSeeRri6maRz1aRixJ989a/bX/g3d+K03i79kbU/h7d3fmS+FvEUqQpu+5BModR/wB9B/zqKOI9pOx5eR8TYjM8wdOa93ofoOmAcCnVHG2WqSu0+6CiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAJwM1yfxf1QaN8Oda1TP/AB76fK/5LXWHpXn37SDsnwV8RMOp0uYf+O14+fVPZZPWmukX+RFT+Gz4QRDnLtupw2d6anNIXcd6/hutrXk33Z883eTEaOM8E8EkVe8K+Ktd8F6tHrnh7UJLeaFsqY2xn298966Dwh8JLv4j+HrjUfA14JtQsUDXelyMA5XpuQ9xkH9K5S9sbzTbuXTdUtWhuIXKyRyKQyt716UcLmWW8mLppxT1TRfLOnaSPs79nb9ovSfitpy6Zqcsdvq8KjzYS3Eox95f616zHL2avzh8NeIdX8K67a69oNy0VzBIGj8tsZ56fjX3z4Z8UN/wr618V+JmjtW/s8T3jyNhYwFyzEnoAK/pXw34wrcQ4N0K+s6aWvc9bD4jnp+90NPxh4x8N+CfD914o8U6zb2On2MDTXV5cyBY4kAyST9K/HX/AIKUf8FufFvxTvb/AOD37LupS6X4fDPBea/G22e8HIOzuin8643/AIK6/wDBUzxD+054yuvgb8H9YktfAel3DR3cttIVbV5lb7zH/nmCMqvfGa+Fo4wB+h21+h18Q3eMD884i4oqVG8PhXZdWS3l7e6pcvqGpXUk80jFpZpH3Mx9yetI9td29tDdT2kscVwGNvK8ZCygMVJU98EEcdxTJAEXgfSvvr48fsn6VpP/AARW+G3xdv8ATY01zS9Ue7W52/O1rfTs3lk9cAlCPTn1rjjB1LtdD47D4Orjo1Jq/uq7b6nwLID5fX8a/Tz/AINr/HEVn8QfiB4Ce4wLrTbe7jj/ALzK2CfyavzDVg67D/8Arr7w/wCDe7UJrP8AbYu7GN2WO48M3IkXs2MEVrh/4iPR4aqunnELdXY/cmFskAVLVeJhtwBU6HI617J+3LYWiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigBGOFJrhP2g7Vr34O+ILdB839mS4H/ATXeEZGKxfF+lpq/h280yQBhPbSR7SOuQa83OKLxOV1aS6xf5EzXNFo/OlCBHg9aTORU+t6fLouq3WlTjbJb3DRP+BxUSbTzX8M4unKjipwfRs+dlH3rM3fhH49vvhv8QNP8UWUh8uOby7qMf8ALSI8Mp9eP1r6Y/aF/Z+0T4seGP8AhNvCUEceprb+bG0a8XCYyA2O/oe1fI0h2j5B7/jX3B+yz4kfxR8GdKkuJd0lvGbeT/gHHNfrXhp9XzzD1snxeqauvJ+R34NqpF05HyV8IvA154t+Kem+FLuFoyt5/pKsv3Qhyc/lXJ/8F2/267j4N/Dq2/ZM+GGrmDWfENqG16e3kIa2scYWPORzJg5/2frX2lrXw98JfB7xV4j+P9+8cNra6PLPMv8AzzKgs7fiAK/nl/au+OGu/tIftBeJvi9rt20z6lqUjW+5s7IQSEUegAr9T4O4X/1WwtWDXvSk1f8AurY+e4lxv9l5f7KL96XXyPO4lJO6Qd8/mf8AJqYY2037q9abI20Z719X10Pyhy5n6nUfBT4Y638aPizoPwt0K3kkuNb1SG1UImSisw3N+Aya/XL/AILXSaB8CP8Agmx4c+COmqkX2jUbCwtYV7rAm5jj/gNeQ/8ABv3+xJdar4hvP2v/AB9pPl2tmrWvhhbiPHmydHnGewHyg/WvLf8AgvF+1jY/G/8AaVtvg94Vv/O0TwLE1vJJG+Umvn5lI/3QFT8DXYkqeHb7n2uForLeH51anxVNEfCsJCjcR0r7m/4N/wCCS5/bm8xWwI/Dd0W/IV8MOxHIr9N/+Dbv4PXOp/EXxp8bLuyk+zafYJp1nOyna0rtlwD3IUfqKyw9/bI8jhujKtm1Pl6M/X6AYOT6VYQYGaZCgPNPTpXsn7f0HUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFV51Dgjb1/WrFMkwDzUzjzRswPiL9rn4czeCvihPq0EO201T99G38If+IV5fFIMV9zftCfCG1+LvgebSkRVvYMy2MxXow/h+h718O6tpWoaBqtxo2r27Q3FtKUmjYfdYV/JfiVwzWyfOJYmnH93U19H1R4+MounUv3I5DGqZFfU37A2qST+C9W0x5Plt9QDr7Bk/8ArV8sYG3k9q+j/wDgn7MVPiG2J4zA2P8Avus/C2tKlxXTj/MmvwJwb/fo5T/gt38bJfhD+wxr1nYXbQ3viS4i0y38t8MVc5c/TaOfrX4GQgLgY2/1r9Y/+DlbxzPDonw5+HMU37uW6ub+SPPXChBX5OtKFXk/hX9R4uV6lj854wryrZo4fyoWZgV2p1xxxXvH/BPf9hvxx+3B8bLXwnpFlLH4f0+RJvEGrbTst4M/cBx99ugH41H+w5/wT++NX7c/j9dD8D6RLa+H7WZf7a8SXEJ+z2yH+FD0d8DhAe3OAa/e79kz9k74W/sg/CXT/hV8LdFjhgt13X186AzX1xj5ppGxySfwA4HAp4fD80k5bFcOcO1MdUVasvcX4nn/AO0rB8TP2a/2XLP4OfsbfCS41TWmsRpmjJZbI4dPXbtM7sxHPfjJJ5r8zPCv/BA39ub4l6vc6/8AETXPD+j3F5M001xdXzXDSMx3HJVc559a/cXyAVy8dPSJdu3bxXoToxlvsfoGNyDC46cfa35Y7LoflV8O/wDg2w0QRx3XxP8Aj9eTNwZrfSdPCKD3wzk/yr9A/wBkP9lH4Y/sd/CW1+D/AMLbGRbKCRp7q6uGDTXcz/ekkIA3HAA7cKK9V8sLxto2hThVqqdOnDVHRgcnwOX+9Rjr3CLK8GnJ0oUdytOrS9z1gooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACgnAzRTZAShAFADXuFXtR9pjzjPbNfGP/AAVX/wCCh/xT/YS8WfCbQfh34L0PVo/H/iRtO1JtaWXdboHhG6Py3Xn94fvZFfVniPx54X8D+Cbnxv4212z0zTbKz+06hfXlwIooI8ZJZjgAUAdF58ecE05JUc4U18DfDz/gpp+0h+2z+0Y/gX/gn98FdHvfhtotwYfEnxQ8aQ3K2srhuVtY43QyYAOOSSTyFAzX3bpYuYLWP+0HVpljAkdE2qWxyQMnA9sn60AXqKha5APBFCXJYbsUATUUzzgFyTTRcZ7UAS0VH9oAOKI7hZBu7UASUZqJ7kKeKPtA6nFAEtFQrdZfaFqTzR2oAdRUbXKL1pr3KKc8UATUVGkyld2felEyZ2g0APpGYAZqI3QXklab9oEg+XHNAEnnrjmgzwsPv18M+J/23Pj5pn/Bbnw7+xDa69Y/8K/1HwZNqVzZtpcf2g3C2ksoYTffA3IvGccVpf8ABSef/grdb/FbwL/w7+k0ePwr5q/8JQLi3tpJXm8zpN56nbBs7x4bOcnpQB9oSLv4H0xXh37Tn7Mlv8QbVvF/hGGOHWIU/eRdFulHY+h9DXsnhg64NFtG8Si3XUPssf277Nny/O2jftzzt3Zx7VZneIfK/rj614ueZHgc9wMsNiI3T28n3IqQVSPKz83tS0+/0m7m0rVrOS3uIHKyQyggqRX0N+wCP9O8QSA/L5UA/wDQq9P+OH7NXg74vW/27/jw1eNP3GoQqMkD+F1/iX6YI7EdK579kn4R+LvhPd+ILHxXZLH5s0It54zlJVG7LKf8a/FuH+A8w4b4ypVLXo62l/mcNLCyo1tFofnl/wAHD2ieM/iP+0p8Ofh74H8Nahq2oSaDObey021aWR2aYdFUE/4VkfsLf8ECPiB8Qb2z8d/tc3r6Dou4S/8ACMWMg+23IxwsjjiJfXGW9NvUfrrr2m+B9Cu5vH2uadp0E1ralZtUuIUWRIRzgyEZ2g9s4rwnVf23vFfj7xHL4V/Zc+Dtx4maFik+tXjNHZofXA6j6ketfvLw9OpU55M8iXDOFrZlLFVfeu72PbfhR8IPhz8EvB1j8P8A4X+ErPRdHsIwlvY2EARU9zjqx5yxJJzznJNdWrIAFBr5xXxF/wAFK5ofto8DeAV7ratJJnHpxNUNp+2n8RfhjrEOh/tQ/BebQYZmCx65pcjTWpY8ZYHO0fQmurbQ+lp06dKCjBWXkfTCFduKUn5eKy/DvibRfFOjW2v+H9QhurO6jWS3nhbcrKRwa0DMo6CgsehPenVGLhcZyKckqscZoAdRUbTBTjNDXCr3FAElFQrc7uak81SOG5oAdRUBuwGwR7UoucjJFAE1FRG5GMinJLuXcaAHk4GaaZQOopjXC9Nw968H/bO+N/xC+DmseB7TwJqMEMeva4bXUPPtlkzHuiGBnofmPNAHvInQ9KUTxk4zzVK8JS0kbPIjJ6e2a8X/AGOPjb4++MMvi1PHF5bzf2RrklrZ+TbiPbGGYAHHXpQB7oJ16U9XDDivDv2n3/a1TXtD/wCGdWs1sd2dSEkMTMzZ+63mA4THdeevtXsWjTalHpVv/bIh+1+QpuvJzs8zHzY9s5oA0KKhNzgc9acs4xg0ASZopu/PIFNaYA7VIz70ASUZ7VD9sTGQRzQbgjnAoAmopqyZGSKcDnmgAooooAKKKKACiiigAooooAKKKRjhc0Afk5/wc6eK7DwPq/wA8aanFJJa6T4subu4jhUFmSNrdiFB6nA4qx8M/hN+07/wXS1LTPit+0Brt58P/wBnWwnWXQ/Auk3QN54mZOPMuXU7Qp6ZwQo4QZJem/8ABy3oWleKfFP7PPhjW7Vbiy1HxpPb3cDMV8yN3tlZcjkcHqOazdW8G/tRf8EIfiL/AMJ98MLbVPiD+zT4gnV9a8PyMZrvwuzkfvYzzgDpkja4ODh/moHofqJ8KfhV8Ofgx4M0/wCHfwu8G2Oh6HptusVlp+mwhI41Ax26k9STyTyeTVX48/GvwB+zz8KdY+MHxK1lbDRdDtWnvZm6kAcKB3YngD1qj+zz+0V8KP2mPhrp3xX+Dvi611jSNSgEkckD/NGxGTG69Uceh6V8Y/8AByZF43vP2AIdN8PTrBpt14y0+PXrhif3cJY7Scfw7uW9hQIwfhv+0H/wV0/4Kbed8Qv2W9f8O/A34X/aGGh6/rWkpfapqsYYgyCOVJEAOAR8ij/aNb3jP4Rf8F2f2aLOT4leC/2t/CfxstLGPzLzwbr3hW30+a6UYLLE9vFGd2AcZfk8bT0riPgR+xd/wW3t/hF4ch+FH/BS7wTZ+Hf7HtzpFrb+EbVljtyg2DP2A54xyTn3NdbJ+x1/wX4BG7/gp/4P4/6k+09f+wfQB9F/8E8/+ChPgT9vH4d32qWeiXHh7xb4dujZeLvCOoMPtGnXIJHtuQkHDYHp1r5a/wCCi3/BVv47/sYf8FIvCvwb0eJtX8Fah4cW4uPDNhpMUl3f3jhljjSUqXXL7R6Cui/4Ju/8E1v2p/2Yv2x/F37Tnxo/aO8H+LrjxZp8lv4ot/D9o1vJLcZUrI0SQxxKQygngE5J5JzXmv7YnhbSPFv/AAcQfB3Ttatkmhj8NxTbGUMN6JIynB9CKAPXbL4Wf8Fuf2nbFfiNqv7U3hT4H2F5GZNP8HaN4Wg1S5gjJO37TNco/wC8xjOwgdeBXOfCz9uH9ur9h39rvwr+x9/wUg1TQfGGg+PJGh8F/FDQbFbVpLgFVMU0SBIwdzKpARSM5ywOR+jSRlvlC1+b/wDwcOLDpUH7O/iaGBVvLX4zWoimx8wUxliM9eqr+QoA+kv+Cq37RHxP/Za/Yj8XfG34O6tDZ+INJt42srie1SdVJkA5VwQev0r5d/Zr/bg/4KXf8FIfhPoVh+ylFofgyytNMjh8XfFjxRpIk+0ahj97HY2u0xnb3ZlZfYV7J/wXMJP/AAS48cSO3LWNu3X/AG1r0T/gkv4M0TwZ/wAE7fhPpegWMcML+FYZ5PLTG+RyWZj7kmgD5h+Omh/8Fsf2E/B15+0g37Wvhv4z+HvD0JvPEfhXVPCNtYS/ZV5lkje3iVhtUE5DcAZ2sBg/Z37FP7Wfgf8AbU/Zo8MftG/D6CS3s9fsybjT7jHmWV1G3lz2744JRwwyPvAAjgium/aT0201H9njx5YX8CyQy+DdUjkjboyG1kBB9sV8O/8ABt5eX0v/AATP1KK0Y+Zb+OtYW13HhTshYfrz6cmgDpv2vf8Agpp8cNY/aKk/Ye/4J3fD7T/FfxBt1B8Ra/qx3aboCEfx4YbnAOcEgDpgniqafsd/8FwbzSv7cuP+CpXh231dvmOlQ/D+yNkp67C5td5A9SCfevh//gmN+z9/wUY+Lvxg+NPjP9mn9rDwz8PfFUfja6g8XWutaLFdXlxIZ5GLBpLaUiMMCOCO1fZC/scf8F+V+Uf8FP8Awf8A+Efaf/K6gDc+AX/BSL9pv4DftHaN+xx/wU78D6LpOseJJRB4N+Inh/KabrMhbasbqSRG7nbyCBkgFV4r7M+L/wAYPAfwK+GOr/Fv4m+IIdM0HQ7Frq/vJm+6ijJGO7HgAdya/MP9qL/gkf8A8FW/2s9J0bwp+0d/wUG8B6yumaibrRBLoaWs0NyBwyPBaRvnj7oPUA44r1v/AIOGYfG9r/wTStbK4vGYDxDpSeKprHO14hnzGBIztMgXqO4zQBheCP2p/wDgqn/wU51O48R/sX/2D8G/hZDcPHpvjTxJpq3mo6oqnBeOKRHRVI5A2D/e7V2WpfAr/guN8ArZ/HvhL9szwf8AF9bOPzLrwf4k8I2+nG8UcssU1tGjK+M4JfAPY9D9U/scxfDq1/Zb+H8fwnFuNBHhGwOn/ZcbNvkLknGPmznJ9c16a2NmQM45FAH4z/svftQXX7WH/BwZ4N+IOs+ANQ8J65Y+B7/TPEXh3UlxJZX0NhcCRAf40JIwfTFfW/8AwV7/AG2P2hP2RviD8EPD/wAEvElnp9p428WtYeIFutLjuDLAGhGF3g7DhzyPWvEtUh8DQf8ABzn4cbweIRcSeA7ttYWBV2/aP7OmyTj+Irtznmtb/g4WK/8AC2/2ZQD/AMz9J/6HbUAffn7SP7RPw0/ZV+DmsfGz4s62tlo+j2plmbjfM/8ADEg/idjwBXwv8OvjV/wWD/4KTj/hY37P+v8Ah/4B/DWeUnRNS1bSY7/VtThzjzNkscigY5GFUf7Rp/8Awchi5PwI+F8GrmYeF5PiRajxN5edph4xu9uuK+/Pg7b+EU+F3h+PwGtv/Yv9k2/9mi1x5fleWu3GPagD4q1n4Zf8Fvf2XYJPiRov7TXhH476bZor6h4N1nwzDpd7NGG+Y201tGn7wL03s3sprh/+Cf8A/wAFaPjh+2b/AMFPNa+C8ttLovgS18KyT/8ACLappMcd9YahFHGJkeXAchZN+Ox9B0r9MpAqRNuX1FflX+zbB4Nt/wDg5G+I48Ex2ywN4Qka6+ykbPtP2eLzc4/i3dffNLlT6AfVn7d/iLxB49+JPhP9mbw5qklrDrU6z6s0bbSyZ4Un0xk4719EfDD4aeFPhd4Ws/CPhDRYbW1towv7tQpc/wB4n+IkjJJr5x/bYt7v4W/tEeBf2hXtZJNPtpVtdQkx8sXPBP4E19Q+GNe0nxHo9vrui3sc1vdQq8UiNkMpGadgNQx4U4FZPirwV4a8caJN4d8XaJBf2NyMTW9wuVYf0/CtcuuOGqnf6tZaZHJd392kMMa5kkkYKqj1JNAFHQfDPhzwJocOh+HNKt9P02zj2w21uoSONf8APc14H4p/ac+Lvxg8f3nwu/ZV0O1kj0+TZq3irUl3W8B6YjXox9yevQYr2L4x6lc3nwg1688M3fmSvpEzW8luwbOU6j8K8r/4Ju2+gxfs828tmsRvpNSujqmMb/N8xsbv+AgYzzigBtz8Gf26bCzTVLL9qLTbq8X55LK48PwRwyEfw7li3YI78fWrnwO/ap8V3fj1/gj+0B4ah0PxVGv+hzQZ+zX6junXBxz1OfavepTG0ZCmvln9vVLQfEf4b3OjHbrq68oh2H5jHkZHHJGcUAd/+2x8YPHPwX+EsfjDwJqEdtefb0iaSa3WQbT1GGrlPC/xP/ah/ab0mPUvhDPY+EfD/lqq+INUs/OuL2QcM0URBVVz6g5x2qT/AIKTRtL+ztbpL95tShHIr234R6NYaH8OtE0jToFjgt9KgSONBgKuwdqBHz54+1j9sv8AZe09fiH4k8f6f468N27j+17RtNjt54EJwZFMarjr1+bHcYr3jQfip4a8RfCeP4taXIZNPl0k3qhV+bAUkqR6ggisv9qiGGT9n3xckkQZV0O4PzD/AGDXJ/sM28Gr/sleH9O1CLzIXt7iGRX5DqZX4P4HFAzg/h1rn7X37Umiz/E/wd8XdN8HaRJcyJpWlQ6bHcOyq2MyM6Mfrj8BXsHwEv8A9oJ7C80n48aZpi3FnceXaapprYF6n98p/D+gOegxXlt7+zr8dv2dNYu/Ef7L/ieLUdFuJWnuPCGscqGJyRG3H4cg/WvSP2b/ANo+y+O1hqFhf+HptH17RZvI1jSbg5aKTOMj1XI9KAML4EfGH4g+Of2jviH8P/EeoQyaX4fmjXTYUt1Vkz1yQMn8c17eFZVxivmf9maY2X7aPxcsZm+eWSGRFbqRwc/qPzr6YEoKfpQB4X40+NPxA0j9sjQPg5Y6lGug32nNNdWzW6li21j94jI6CuM/4KPk/wBvfC8/9TQf/Qoan8cyNf8A/BSDw1HbLu+z6HI82P4Rsbr+dV/+CkB3a/8AC8lf+ZoI/wDH4aAPZPjr4e+PWu2FqvwU8baTo3liQ6g2qWhlEq44C4U4xzXyz+xx4V/ah1a58VH4XfEfQdOEevsusfbrEyefMGOWT5DtHX0r7kv9p0+bj/lm38q+a/8AgnRgTfEEgf8AM0TY/wC+2oA3P2uPjP8AFH4NyeA7Pwvq8EUutat9l1ZmtVkEoHl5xuHy9T0rpP2xfid4x+D37PupeP8AwNfR2+pW0lqI5ZrdZFAeVVbKsCOh9K83/wCCizj+2fhnk/8AMyt/OKuq/wCCiah/2S9XXb/y8WJOf+u6ZoA57wV8W/2oP2mPDtrL8InsPDOkpaxre+KNVs/Mlup8fP5ERG0LuyMkEccYqHx6/wC2v+zrpTfEnUfifpvjjRrM7tU02bS47eVI+7qY0Xp9Tj0r2n9nTw9p/h/4IeF9O0y3CRrosB+VcZJUMT9STU3x/to5Pgr4mEqKy/2NcFlYf7BoAs/Cn4kaV8VPAOn+O9EY/Z9Qt1kVD1Qkcg/Q1498R/2p/iL4v+I1z8Ff2X/DNtqmrWQ/4m2uX3zWtic46ZG4jnqQM9N1O/YVn1F/2O4jph/0pY70Wv8Av5fb+tZf/BNGLST8MfEX2vyz4g/4Sq4OsFv9dnbHs3d9v3sds5oA2Lb4K/twrD/a1x+1Fpf2zr9jHh2Brc/7O7yw2PcDNcH8bf2vP2ifg9osPgjxx4bttK8TSX0YtdYs4BNY38G7DEBuUf25/Cvr9dgXla+Zv+Cl/wDwjv8AwqzSBqHk/bv7ehNjn73X5se2KAPorwne3Wo+G7C+vCGlmtI5JCB1YqCa1Kx/Ap/4o/SiT/y4R/8AoArYoAKKKKACiiigAooooAKKKKACkf7tLSPnbwKAPz//AOC0/wCxp8fP2sPG3wS1T4LeEP7Vg8I+Lnvdcb7QqeRDvgO75iM8I3A6fjX3Ff8AhbSPE/hR/CviTSbe8s7q0FveWtxGGSRCuCpB6jGa1/JBb5o+alCMvAFAH5peJv8Agnx+1p/wT0/abh+OX/BNKMa14D8RX2fGXwqvrwLDCCcs9uWI2jrtxgr0O5Tx9ufFj4P+Ff2uP2ebz4W/GjwZJb2PiTTBHqGl3W0y2chHUMMjcrcgj0r1Ax4OVSgRAdvegD80/hH4U/4Kv/8ABMG3uPhN4Q+Fdn8dvhlZyEeG5LfU/s2q6fDniMgq24AYGMYHZu1b3jT9sn/gsL8eNNbwB8C/+Cf3/Cv7y8Xy28XeMNYDx2YPBZYtigsAcg7uo6HpX6GGEbtxjzTlhAGNo49qAPmH/gm3+wVrn7GXgjWNb+KXxNv/ABl4/wDGF99u8V65cXUjQCQ8+VCrk4UEnLYBY44AAryP48fse/Hbxj/wWh+HP7V+geE/O8FaF4f+zalqn2hR5cmxxjb17ivvspjgCm+QofcFoAISoGR6V8Tf8FpP2Tfjd+1doXwhsPgr4X/tSTwv8TLfVtZXzlTybVYnBfkjOCR719uBcNyKbLCrtuK0AfL3/BVH4D/Ev9pH9gnxR8GvhPorah4g1KzgS0tfMC7mDKTycYr0j9g74deLPg/+x/8AD34YeO9O+yaxofhq3tNQt9wby5VByM969X+zqeCvfIp8UQQY20Ac18adC1LxV8I/FPhjRoPNu9S8N3trax5HzSSQOijn/aIr5R/4Idfsr/Gb9kL9ji6+E3x18NHS9ck8Z39+tsJ1k3QSJCEbKkj+FuOvFfaU0Rk47UJbxpyE5oA+Af2sf+Cdv7Sfwr/aYn/bp/4Jva/ptj4m1JceMvBGq/LZa5jGZOoAc456HPIPUFzf8FHf+CnWmWP9j3//AASd1yXV448NJb+JFNm7DgsH8nIXv0NffjRbsjGM0xoNxwUz+FAH53/Bv9ij9vH9sH9pfw/+1T/wUN8VWvhfQ/CN4Lvwp8MfC984TzQQUa4dT82CAWyfmIxhBkH7f+NvwT8B/tC/CjWPg78UfD8WpaHrlk9tqFrJxlSOCp4wQcEHIwRmu0WJdvK1JgY5FAH5h/C/4Lf8FT/+CVF/dfD34CeDLT44fCRZ3fRdJnvjbappkbHPlq3zdB2ClSeRtziu81X9sL/grn8dLGTwJ8HP2BoPh9fXi+XJ4q8ca1vgslIwXWEKvmMByPm6/wAJ6V9+GEZ4WjycjDLQB+U/7In/AASq/ah/Zs/4KzeHf2kfHniW88Z6ZN4dvZfFnjK6lVfM1K4tJVZET7wjDFVAP14HFezf8Fi/2O/jz+1J8Q/gbr/wc8Jf2lbeDvFzX2uv5yJ5EJaE7sE88I3T0r7ua3BfmPP4U8Q8cjt+VAHnH7UP7NPw3/a1+C+r/A/4taP9q0nWLUoWH+st5R9yaM9nVuQfwr4V+E3hz/grf/wTCtn+Evh34XWvx6+G1pJt8P3FpqH2bU7GHPCNkMWAH8JGPR8cV+mvYZWo3gD8FfbpQB8Aa9+09/wV9/ab0mT4d/Bj9jCz+EzX6+XdeNPG2sCY2UbcO0UCou58ZIbLYI5U15z/AME6/wDglt+0J+x5/wAFQdZ+LXizUb7xJ4YvPC0qXHjTUZV8y/1GeNGlOwHKrv3gdeAMnsP1E+zrgDYPyoSAZU7OlAGH8RPh34X+Jnha68H+MNJS6s7uPEiMOV9CD2Ir5usPgJ+1p+zHqEi/AjxNa+JPDbSFo9F1T78ak/dU5BB9w2PavrMjIxTNuRyP/r0AfNw/aX/a6j/0R/2Rbg3H3d66qdufX/VdDWPqvwd/a7/advo7T4w67beD/DPmBp9J0t9006/3Sc56euB/s19UGFS2dn405Y9tAHO+CPh/oXgTwha+CdDtWWwtLfyo45HLlhjnJPUn8q8I1z9nz40fs/8AxBvPiN+zLNa32j6pL5uq+EtQk2qH7+W3v+Y6YIFfTeB6U14lZcYoA+d2/aO/aou4f7N039k68ivnXHnXOoAQo394nZk4/DPtUnwZ/Zm+IWq/EY/Hj9pDW4dQ8QBcaXpdqP8AR9PHt1BPp+ZJNfQXlZ+8O9OVCvQUAeI/twfCvxr8W/hFD4Z8EaZ9qvF1GOUx+YFwo6nJr1nwTaXGneGdP0+7j2yw2UaSD0YKARWqyBlwVzSJGUOaAOM+PfhzWPF/wk8SeGNCtvOvL7SZobaLdjc5UgDNcX+z38PPif8ADn9lqz8DW1vb2PiS1trhYFu/mjSRnYqWweRzmvZ5I88haRIiB0oA+cdO+N37YnhbTh4b8Wfs5/2xqaDYuqadfBLeb0cjadv4E/hW1+yd8DPiB4K1bxH8WPirJAviDxXeedc2Nr9y3TJIX0zz6npXuixYOMfjinKgWgD53+O3wA+Keg/F6P8AaL/Z6ubZtYMAg1jR7xsR3yDjPPcjGeQcgYps37R37VU9r/Ztr+yfeJflcebLqX7hW/vbtuSPwr6KZATytN2dylAHgv7OH7PnxE0jx9qXx6+OF/b3HibVYfKhtbcBo7KLrtBqv+2t8GvH/wAVtW8CXXgnRjdLouvfadQYsBsj3R88n/ZNfQSxFegoZMjBWgCpdB2spIlG7dGwH5V4h+xd8IPHPwqn8YN400n7KNU16S5svnB3xlmIPHTqK95ETA8rx2pfIAHHX1oA8A/bT+Dnjz4ran4HufBekNdro+uNc3+HA2JlOeevQ1v/ALZXw48W/FH9njUvBPgywa61G4mtTHFuAyElRm6n0B/KvYBAM9KPKG3GKAOf+EukX/h74caHoWqQeXc2elQQzx/3WVACP0qL4waPqXiL4b654e0iHzLq802aK3TplipArp0UKOlNeIMxO2gDyD9jL4c+KfhP8CLLwd42077LfQ3EzTR7g2FZyR09q4nxh8ANfi+Id78a/wBkX4h6dZ6pK/8AxPdDkk32t2/XnB+Vic8HvyCK+kXtsxsqx/eGDnuK+efG/wCyt8T/AIffEC8+K37MXjGCwuNSbfqvh7Ukza3DcncvPGSehxjJwcHFAFq2+KX7bzRNpUn7P+i/atu0339uhYd397bt6e27PvXi/wC1r8IviLBoGk+Pvjj46j1DxJqGtW9tpejaWClrZRlssFGSXPv+pr2RvHf7d0enLYH4K+H2vAu1r0awPL+oX0/H8qh8C/sv/FPx34/s/ix+074qt765sG36XoOnA/Z7V/U+pH1Oe5oBo948HW8kHhewgl+9Haxqf++RWtUNvGbeIRqnA44qYUAFFFFABRRRQAUUUUAFFFFABRRRQAfhRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUZPpRRQAwIc89qdj3paKAAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQB//9k="
    img = XLImage(io.BytesIO(base64.b64decode(LOGO_B64)))
    
    # Scale: Adjust these to fit your logo's aspect ratio
    img.width = 240
    img.height = 70

    # We anchor it to C1
    # Col 2 = Column C
    # colOff: We move it right by roughly half the width of Column C to center it
    # rowOff: We move it down slightly from the very top (Row 1 is 90 height)
    
    c2e = pixels_to_EMU
    h_offset = c2e(120) # Push right from left edge of Col C
    v_offset = c2e(10)  # Push down from top of Row 1

    marker = AnchorMarker(col=2, colOff=h_offset, row=0, rowOff=v_offset)
    size = XDRPositiveSize2D(c2e(img.width), c2e(img.height))
    
    img.anchor = OneCellAnchor(_offset=marker, _size=size)
    
    ws.add_image(img)
except Exception as e:
    print(f"Logo placement error: {e}")

wb.save(out_path)
